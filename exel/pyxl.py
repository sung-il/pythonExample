# import openpyxl
from openpyxl import load_workbook
 
wb = load_workbook('score.xlsx')

ws = wb.active
# ws = wb.get_sheet_by_name("Sheet1")

for r in ws.rows:
    row_index = r[0].row
    kor = r[1].value
    eng = r[2].value
    math = r[3].value
    sum = kor + eng + math
 
    ws.cell(row=row_index, column=5).value = sum
 
    print(kor, eng, math, sum)
 
wb.save("score2.xlsx")
wb.close()